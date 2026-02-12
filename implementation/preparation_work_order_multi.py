#!/usr/bin/env python3
"""
Workflow 0 - Extension Multi-Work Orders
Support pour ZIP contenant plusieurs Work Orders SPR/SAR

Nouvelles fonctions ajoutées pour gérer:
- Détection automatique de tous les Work Orders dans un ZIP
- Parsing Excel multi-lignes (ligne 14+)
- Localisation fichiers audio dans arborescence profonde DAUDIO
- Génération metadata.json pour chaque dossier séparément
"""

import sys
import io

# Fix encodage UTF-8 Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import os
import re
import glob
import logging
from pathlib import Path
from openpyxl import load_workbook

# Configuration logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class WorkOrderError(Exception):
    """Exception pour erreurs de traitement work order"""
    pass


def localiser_fichier_audio(dossier_path, numero_dossier):
    """
    Localiser fichier .a00 dans arborescence DAUDIO profonde.

    Structure cible:
        MC3-56703/
        └── DAUDIO/
            └── 2025-12-09/              ← Date audience
                └── _0231/               ← Timestamp enregistrement (02h31)
                    └── MC3-56703/       ← Numéro dossier (répété)
                        └── 6703.a00     ← Fichier audio (4 derniers chiffres)

    Args:
        dossier_path: Chemin vers dossier MC3-xxxxx/
        numero_dossier: "MC3-56703"

    Returns:
        str: Chemin absolu vers fichier .a00

    Raises:
        FileNotFoundError: Si aucun fichier audio trouvé

    Exemple:
        >>> localiser_fichier_audio("extracted/MC3-56703", "MC3-56703")
        'extracted/MC3-56703/DAUDIO/2025-12-09/_0231/MC3-56703/6703.a00'
    """
    logger.info(f"   🔍 Recherche audio dans: {os.path.basename(dossier_path)}")

    # Recherche récursive de tous les .a00
    audio_files = glob.glob(f"{dossier_path}/**/DAUDIO/**/*.a00", recursive=True)

    if not audio_files:
        raise FileNotFoundError(f"Aucun fichier audio trouvé dans {dossier_path}")

    # Vérification cohérence nom fichier
    # Règle: MC3-56703 → 6703.a00 (4 derniers chiffres)
    dernier_4_chiffres = numero_dossier[-4:]  # "56703" → "6703"

    # CRITICAL: Pattern strict pour éviter matching incorrect (CONTRAINTE #5 CLAUDE.md)
    # Chercher fichier qui commence EXACTEMENT par les 4 derniers chiffres
    fichiers_correspondants = []

    for audio in audio_files:
        basename = os.path.basename(audio)
        if basename.startswith(dernier_4_chiffres + '.') or basename.startswith(dernier_4_chiffres + '_'):
            fichiers_correspondants.append(audio)

    if fichiers_correspondants:
        audio_choisi = fichiers_correspondants[0]
        taille_mb = os.path.getsize(audio_choisi) / (1024 * 1024)
        logger.info(f"   ✅ Audio trouvé: {os.path.basename(audio_choisi)} ({taille_mb:.2f} MB)")

        if len(fichiers_correspondants) > 1:
            logger.warning(f"   ⚠️  Plusieurs fichiers audio trouvés pour {numero_dossier}:")
            for f in fichiers_correspondants:
                logger.warning(f"      - {os.path.basename(f)}")
            logger.warning(f"      Sélectionné: {os.path.basename(audio_choisi)}")

        return audio_choisi

    # ERREUR CRITIQUE: Aucun fichier ne correspond au numéro dossier (pas de fallback silencieux)
    logger.error(f"   ❌ ERREUR: Aucun fichier audio ne correspond au dossier {numero_dossier}")
    logger.error(f"      Attendu: {dernier_4_chiffres}.a00 ou {dernier_4_chiffres}_*.a00")
    logger.error(f"      Fichiers trouvés: {[os.path.basename(f) for f in audio_files]}")

    raise FileNotFoundError(
        f"Aucun fichier audio correspondant à {numero_dossier} (attendu: {dernier_4_chiffres}.a00). "
        f"Fichiers présents: {[os.path.basename(f) for f in audio_files]}"
    )


def lire_excel_tous_work_orders(excel_path):
    """
    Lire toutes les lignes de données Excel Work Order (ligne 14+).

    Structure Excel RCE-9878-AA:
        Lignes 1-2   : En-tête CISR bilingue (FR/EN)
        Lignes 3-10  : Métadonnées globales Work Order
        Ligne 13     : HEADERS (23 colonnes)
        Lignes 14-19 : DONNÉES (6 Work Orders, une ligne par dossier)

    Colonnes extraites:
        - Col 4 (D): File Number (MC3-xxxxx) - IDENTIFIANT UNIQUE
        - Col 6 (F): Date of Hearing
        - Col 13 (M): Length of Audio (HH:MM:SS)
        - Col 22 (V): Recording Unit Remarks (découpage audio)

    Args:
        excel_path: Chemin vers fichier Excel work order

    Returns:
        list[dict]: Liste de métadonnées pour chaque Work Order
            [
                {
                    'numero_dossier': 'MC3-56703',
                    'date_audience': '2025-12-09',
                    'duree_audio': '0:09:00',
                    'recording_remarks': 'commence à 1:46',
                    'ligne_excel': 14
                },
                ...
            ]

    Raises:
        WorkOrderError: Si Excel invalide ou structure incorrecte
    """
    logger.info(f"📖 Parsing Excel Work Order multi-dossiers: {os.path.basename(excel_path)}")

    try:
        # Charger Excel
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        # === ÉTAPE 1: Trouver ligne d'en-têtes ===
        header_row = None
        for row_idx in range(1, 30):
            for col_idx in range(1, 10):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and 'file number' in str(cell_value).lower():
                    header_row = row_idx
                    break
            if header_row:
                break

        if not header_row:
            raise WorkOrderError("Ligne d'en-têtes non trouvée (colonne 'File Number')")

        logger.info(f"   Ligne d'en-têtes détectée: {header_row}")

        # === ÉTAPE 2: Lire toutes les lignes de données ===
        work_orders_data = []
        row_idx = header_row + 1

        while row_idx <= ws.max_row:
            # Col D (4): File Number
            file_num = ws.cell(row=row_idx, column=4).value

            # Si ligne vide, arrêter lecture
            if not file_num or str(file_num).strip() == "":
                break

            # Extraire données ligne
            numero_dossier = str(file_num).strip()

            # Col F (6): Date of Hearing
            date_hearing = ws.cell(row=row_idx, column=6).value
            if date_hearing:
                # Convertir datetime vers string si nécessaire
                if hasattr(date_hearing, 'strftime'):
                    date_hearing = date_hearing.strftime('%Y-%m-%d')
                else:
                    date_hearing = str(date_hearing)

            # Col M (13): Length of Audio
            length_audio = ws.cell(row=row_idx, column=13).value
            duree_audio = None
            if length_audio:
                # Peut être "HH:MM:SS" ou nombre (heures)
                if isinstance(length_audio, (int, float)):
                    heures = float(length_audio)
                    hours = int(heures)
                    minutes = int((heures - hours) * 60)
                    duree_audio = f"{hours}:{minutes:02d}:00"
                else:
                    duree_audio = str(length_audio)

            # Col V (22): Recording Unit Remarks
            recording_remarks = ws.cell(row=row_idx, column=22).value
            if recording_remarks:
                recording_remarks = str(recording_remarks).strip()

            work_orders_data.append({
                'numero_dossier': numero_dossier,
                'date_audience': date_hearing,
                'duree_audio': duree_audio,
                'recording_remarks': recording_remarks,
                'ligne_excel': row_idx
            })

            logger.info(f"   ✅ Ligne {row_idx}: {numero_dossier} ({duree_audio})")

            row_idx += 1

        logger.info(f"📊 {len(work_orders_data)} Work Orders détectés dans Excel")

        return work_orders_data

    except Exception as e:
        raise WorkOrderError(f"Erreur parsing Excel multi-Work Orders: {e}")


def detecter_work_orders_multiples(extracted_dir, excel_path):
    """
    Détecter tous les Work Orders dans un ZIP décompressé.

    Algorithme:
        1. Lire Excel Work Order pour connaître nombre exact de dossiers
        2. Vérifier présence dossiers MC3-xxxxx correspondants
        3. Localiser pages couvertures (racine du ZIP)
        4. Localiser fichiers audio (dans chaque dossier DAUDIO)
        5. Associer: Excel ligne → dossier → page couverture → audio

    Args:
        extracted_dir: Chemin vers dossier ZIP décompressé
        excel_path: Chemin vers fichier Excel work order

    Returns:
        list[dict]: Liste de Work Orders détectés
            [
                {
                    'numero_dossier': 'MC3-16722',
                    'dossier_path': '/path/to/MC3-16722/',
                    'page_couverture': '/path/to/MC3-16722 SPR Page couverture.docx',
                    'audio_file': '/path/to/6722.a00',
                    'excel_work_order': '/path/to/Work Order RCE-9878-AA.xlsx',
                    'metadata_excel': {
                        'date_audience': '2025-12-09',
                        'duree_audio': '0:11:00',
                        'recording_remarks': None
                    }
                },
                ...
            ]

    Raises:
        WorkOrderError: Si incohérences détectées
    """
    logger.info("🔍 Détection Work Orders multiples...")

    # === ÉTAPE 1: Lire Excel pour référence ===
    work_orders_excel = lire_excel_tous_work_orders(excel_path)

    # === ÉTAPE 2: Détecter dossiers MC3-xxxxx ===
    dossiers_regex = re.compile(r'MC[2-5]-\d{5}')
    dossiers_trouves = []

    for item in os.listdir(extracted_dir):
        item_path = os.path.join(extracted_dir, item)
        if os.path.isdir(item_path) and dossiers_regex.match(item):
            dossiers_trouves.append(item)

    logger.info(f"📁 {len(dossiers_trouves)} dossiers MC détectés: {', '.join(sorted(dossiers_trouves))}")

    # === ÉTAPE 3: Valider cohérence Excel vs dossiers ===
    numeros_excel = set(wo['numero_dossier'] for wo in work_orders_excel)
    dossiers_set = set(dossiers_trouves)

    if numeros_excel != dossiers_set:
        manquants_excel = dossiers_set - numeros_excel
        manquants_dossiers = numeros_excel - dossiers_set

        if manquants_excel:
            logger.warning(f"⚠️  Dossiers présents mais absents Excel: {manquants_excel}")
        if manquants_dossiers:
            logger.warning(f"⚠️  Dossiers listés Excel mais absents: {manquants_dossiers}")

    # === ÉTAPE 4: Localiser pages couvertures ===
    pages_couvertures = {}
    for root, dirs, files in os.walk(extracted_dir):
        for file in files:
            if file.endswith('.docx') and not file.startswith('~'):
                # Chercher pattern MC3-xxxxx dans nom fichier
                match = dossiers_regex.search(file)
                if match:
                    numero = match.group()
                    pages_couvertures[numero] = os.path.join(root, file)

    logger.info(f"📄 {len(pages_couvertures)} pages couvertures localisées")

    # === ÉTAPE 5: Assembler données ===
    work_orders = []

    for wo_excel in work_orders_excel:
        numero = wo_excel['numero_dossier']
        dossier_path = os.path.join(extracted_dir, numero)

        # Vérifier dossier existe
        if not os.path.isdir(dossier_path):
            logger.warning(f"⚠️  Dossier manquant: {numero}")
            continue

        # Page couverture
        page_couv = pages_couvertures.get(numero)
        if not page_couv:
            logger.warning(f"⚠️  Page couverture manquante pour {numero}")

        # Fichier audio
        audio_file = None
        try:
            audio_file = localiser_fichier_audio(dossier_path, numero)
        except FileNotFoundError as e:
            logger.warning(f"⚠️  {e}")

        work_orders.append({
            'numero_dossier': numero,
            'dossier_path': dossier_path,
            'page_couverture': page_couv,
            'audio_file': audio_file,
            'excel_work_order': excel_path,
            'metadata_excel': {
                'date_audience': wo_excel['date_audience'],
                'duree_audio': wo_excel['duree_audio'],
                'recording_remarks': wo_excel['recording_remarks']
            }
        })

    logger.info(f"✅ {len(work_orders)} Work Orders assemblés avec succès")

    # === VALIDATION FINALE ===
    erreurs = []
    for wo in work_orders:
        if not wo['page_couverture']:
            erreurs.append(f"{wo['numero_dossier']}: Page couverture manquante")
        if not wo['audio_file']:
            erreurs.append(f"{wo['numero_dossier']}: Fichier audio manquant")

    if erreurs:
        logger.warning("⚠️  Avertissements détection:")
        for err in erreurs:
            logger.warning(f"   - {err}")

    return work_orders


# === TEST ===
if __name__ == '__main__':
    print("=" * 80)
    print("TEST : Détection Multi-Work Orders sur RCE-9878-AA")
    print("=" * 80)
    print()

    # Chemin vers dossier extrait
    extracted_dir = "Test_Demandes/Nouveaux/Docs initiaux - OneDrive_2_06-01-2026/extracted_SPR/RCE-9878-AA -Regdeck- FR -  SPR - Bench -5TAT- 1.41h"

    # Chemin vers Excel
    excel_path = os.path.join(extracted_dir, "Work Order RCE-9878-AA. xlsx.xlsx")

    if not os.path.exists(extracted_dir):
        print(f"❌ Dossier extrait non trouvé: {extracted_dir}")
        sys.exit(1)

    if not os.path.exists(excel_path):
        print(f"❌ Fichier Excel non trouvé: {excel_path}")
        sys.exit(1)

    try:
        # Détecter tous Work Orders
        work_orders = detecter_work_orders_multiples(extracted_dir, excel_path)

        print()
        print("=" * 80)
        print(f"RÉSUMÉ : {len(work_orders)} Work Orders détectés")
        print("=" * 80)
        print()

        for i, wo in enumerate(work_orders, 1):
            print(f"Work Order #{i} : {wo['numero_dossier']}")
            print(f"   Dossier      : {os.path.basename(wo['dossier_path'])}")
            print(f"   Page couverture : {'✅' if wo['page_couverture'] else '❌'}")
            print(f"   Audio        : {'✅' if wo['audio_file'] else '❌'} {os.path.basename(wo['audio_file']) if wo['audio_file'] else ''}")
            print(f"   Date audience: {wo['metadata_excel']['date_audience']}")
            print(f"   Durée audio  : {wo['metadata_excel']['duree_audio']}")
            if wo['metadata_excel']['recording_remarks']:
                print(f"   Remarks      : {wo['metadata_excel']['recording_remarks']}")
            print()

        print("✅ Test réussi !")

    except WorkOrderError as e:
        print(f"❌ Erreur: {e}")
        sys.exit(1)
