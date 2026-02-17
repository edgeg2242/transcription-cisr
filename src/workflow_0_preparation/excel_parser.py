"""
Parsing du fichier Excel Work Order CISR.

Extrait les metadonnees de chaque Work Order depuis le fichier Excel
qui est la source de verite du Work Assignment.
"""
import os
import re
import logging
from openpyxl import load_workbook

from src.common.exceptions import WorkOrderError

logger = logging.getLogger(__name__)


def parser_excel_work_order(excel_path, numero_dossier=None):
    """
    Parse le fichier Excel Work Order pour extraire metadonnees critiques.

    Colonnes extraites:
    - Col C: Surname (nom demandeur)
    - Col H: Language of Hearing
    - Col M: Length of Audio
    - Col Q: Word Count
    - Col U: Name of Transcriber
    - Col V: Recording Unit Remarks (instructions decoupage audio)

    Args:
        excel_path: Chemin vers fichier Excel work order
        numero_dossier: Numero de dossier a chercher (ex: "MC3-03924") - optionnel

    Returns:
        dict: Metadonnees enrichies avec section "transcription"
              {
                  "transcription": {
                      "transcripteur": "Samantha Pitt",
                      "agence": "RegDeck",
                      "word_count_attendu": 1335,
                      "duree_audio_heures": 0.00625,
                      "recording_remarks": "commence a 1:46",
                      "audio_decoupage": {
                          "start_time_seconds": 106,
                          "applique": false
                      }
                  },
                  "participants": {
                      "demandeur_nom_famille": "HAMMOUD"
                  }
              }

    Raises:
        WorkOrderError: Si Excel invalide ou colonnes manquantes
    """
    try:
        logger.info(f"Parsing Excel Work Order: {os.path.basename(excel_path)}")

        # Charger Excel
        wb = load_workbook(excel_path, data_only=True)

        # Premiere feuille (Work Order principal)
        ws = wb.active

        # Initialiser metadonnees
        metadata_excel = {
            "transcription": {
                "transcripteur": None,
                "agence": None,
                "word_count_attendu": None,
                "duree_audio_heures": None,
                "recording_remarks": None,
                "audio_decoupage": {
                    "start_time_seconds": None,
                    "applique": False
                }
            },
            "participants": {
                "demandeur_nom_famille": None
            },
            "validation": {
                "langue_audience": None
            }
        }

        # === EXTRACTION COLONNES ===
        # NOTE: Excel utilise indexation 1-based (A=1, B=2, ...)
        # Colonnes cibles:
        # C=3, H=8, M=13, Q=17, U=21, V=22

        # === ETAPE 1: Trouver ligne d'en-tetes ===
        # Chercher ligne contenant "Surname" ou "File Number"
        header_row = None
        for row_idx in range(1, 30):
            for col_idx in range(1, 10):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and 'surname' in str(cell_value).lower():
                    header_row = row_idx
                    break
            if header_row:
                break

        if not header_row:
            raise WorkOrderError("Ligne d'en-tetes non trouvee dans Excel (colonne 'Surname')")

        logger.info(f"   Ligne d'en-tetes detectee: {header_row}")

        # === ETAPE 2: Trouver ligne de donnees ===
        data_row = None

        if numero_dossier:
            # Chercher ligne contenant le numero de dossier (col D)
            logger.info(f"   Recherche dossier: {numero_dossier}")
            for row_idx in range(header_row + 1, ws.max_row + 1):
                file_num = ws.cell(row=row_idx, column=4).value  # Col D
                if file_num and numero_dossier in str(file_num):
                    data_row = row_idx
                    logger.info(f"   Dossier trouve a ligne: {data_row}")
                    break

            if not data_row:
                raise WorkOrderError(f"Dossier {numero_dossier} non trouve dans Excel")
        else:
            # Sans numero de dossier, prendre premiere ligne de donnees
            data_row = header_row + 1
            logger.info(f"   Utilisation premiere ligne de donnees: {data_row}")

        # === COL C: Surname (Nom demandeur) ===
        surname = ws.cell(row=data_row, column=3).value
        if surname:
            surname_clean = str(surname).strip().upper()
            metadata_excel["participants"]["demandeur_nom_famille"] = surname_clean
            logger.info(f"   Col C (Surname): {surname_clean}")

        # === COL H: Language of Hearing ===
        language = ws.cell(row=data_row, column=8).value
        if language:
            language_clean = str(language).strip()
            metadata_excel["validation"]["langue_audience"] = language_clean
            logger.info(f"   Col H (Language): {language_clean}")

        # === COL M: Length of Audio ===
        length_audio = ws.cell(row=data_row, column=13).value
        if length_audio:
            try:
                # Peut etre un nombre (heures) ou texte "1:30:00"
                if isinstance(length_audio, (int, float)):
                    duree_heures = float(length_audio)
                else:
                    # Parser format "HH:MM:SS" ou "MM:SS"
                    parts = str(length_audio).split(':')
                    if len(parts) == 3:
                        h, m, s = map(int, parts)
                        duree_heures = h + m/60 + s/3600
                    elif len(parts) == 2:
                        m, s = map(int, parts)
                        duree_heures = m/60 + s/3600
                    else:
                        duree_heures = None

                if duree_heures:
                    metadata_excel["transcription"]["duree_audio_heures"] = round(duree_heures, 5)
                    logger.info(f"   Col M (Length of Audio): {duree_heures:.4f} h")
            except Exception as e:
                logger.warning(f"   Col M parsing echoue: {e}")

        # === COL Q: Word Count ===
        word_count = ws.cell(row=data_row, column=17).value
        if word_count:
            try:
                wc_int = int(word_count)
                metadata_excel["transcription"]["word_count_attendu"] = wc_int
                logger.info(f"   Col Q (Word Count): {wc_int} mots")
            except Exception as e:
                logger.warning(f"   Col Q parsing echoue: {e}")

        # === COL U: Name of Transcriber ===
        transcriber = ws.cell(row=data_row, column=21).value
        if transcriber:
            transcriber_clean = str(transcriber).strip()
            metadata_excel["transcription"]["transcripteur"] = transcriber_clean

            # Extraire agence (souvent entre parentheses)
            match_agence = re.search(r'\(([^)]+)\)', transcriber_clean)
            if match_agence:
                metadata_excel["transcription"]["agence"] = match_agence.group(1)

            logger.info(f"   Col U (Transcriber): {transcriber_clean}")

        # === COL V: Recording Unit Remarks (CRITIQUE) ===
        remarks = ws.cell(row=data_row, column=22).value
        if remarks:
            remarks_clean = str(remarks).strip()
            metadata_excel["transcription"]["recording_remarks"] = remarks_clean

            # Parser instructions de decoupage
            # Exemples:
            # - "commence a 1:46"
            # - "commence a 0:33"
            # - "arrete a 8:30"
            match_start = re.search(r'commence \u00e0 (\d+):(\d+)', remarks_clean, re.IGNORECASE)
            if match_start:
                minutes = int(match_start.group(1))
                seconds = int(match_start.group(2))
                start_seconds = minutes * 60 + seconds
                metadata_excel["transcription"]["audio_decoupage"]["start_time_seconds"] = start_seconds
                logger.info(f"   Col V (Recording Remarks): {remarks_clean}")
                logger.info(f"      Decoupage detecte: demarrer a {start_seconds}s ({minutes}:{seconds:02d})")
            else:
                logger.info(f"   Col V (Recording Remarks): {remarks_clean}")
                logger.info(f"      Aucune instruction de decoupage detectee")

        logger.info(f"Excel Work Order parse avec succes")

        return metadata_excel

    except FileNotFoundError:
        raise WorkOrderError(f"Fichier Excel non trouve: {excel_path}")
    except Exception as e:
        raise WorkOrderError(f"Erreur parsing Excel: {e}")


def lire_excel_tous_work_orders(excel_path):
    """
    Lire toutes les lignes de donnees Excel Work Order (ligne 14+).

    Structure Excel RCE-9878-AA:
        Lignes 1-2   : En-tete CISR bilingue (FR/EN)
        Lignes 3-10  : Metadonnees globales Work Order
        Ligne 13     : HEADERS (23 colonnes)
        Lignes 14-19 : DONNEES (6 Work Orders, une ligne par dossier)

    Colonnes extraites:
        - Col 4 (D): File Number (MC3-xxxxx) - IDENTIFIANT UNIQUE
        - Col 6 (F): Date of Hearing
        - Col 13 (M): Length of Audio (HH:MM:SS)
        - Col 22 (V): Recording Unit Remarks (decoupage audio)

    Args:
        excel_path: Chemin vers fichier Excel work order

    Returns:
        list[dict]: Liste de metadonnees pour chaque Work Order
            [
                {
                    'numero_dossier': 'MC3-56703',
                    'date_audience': '2025-12-09',
                    'duree_audio': '0:09:00',
                    'recording_remarks': 'commence a 1:46',
                    'ligne_excel': 14
                },
                ...
            ]

    Raises:
        WorkOrderError: Si Excel invalide ou structure incorrecte
    """
    logger.info(f"Parsing Excel Work Order multi-dossiers: {os.path.basename(excel_path)}")

    try:
        # Charger Excel
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        # === ETAPE 1: Trouver ligne d'en-tetes ===
        header_row = None
        for row_idx in range(1, 30):
            for col_idx in range(1, 10):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and 'file number' in str(cell_value).lower():
                    header_row = row_idx
                    break
            if header_row:
                break

        if not header_row:
            raise WorkOrderError("Ligne d'en-tetes non trouvee (colonne 'File Number')")

        logger.info(f"   Ligne d'en-tetes detectee: {header_row}")

        # === ETAPE 2: Lire toutes les lignes de donnees ===
        work_orders_data = []
        row_idx = header_row + 1

        while row_idx <= ws.max_row:
            # Col D (4): File Number
            file_num = ws.cell(row=row_idx, column=4).value

            # Si ligne vide, arreter lecture
            if not file_num or str(file_num).strip() == "":
                break

            # Extraire donnees ligne
            numero_dossier = str(file_num).strip()

            # Col F (6): Date of Hearing
            date_hearing = ws.cell(row=row_idx, column=6).value
            if date_hearing:
                # Convertir datetime vers string si necessaire
                if hasattr(date_hearing, 'strftime'):
                    date_hearing = date_hearing.strftime('%Y-%m-%d')
                else:
                    date_hearing = str(date_hearing)

            # Col M (13): Length of Audio
            length_audio = ws.cell(row=row_idx, column=13).value
            duree_audio = None
            if length_audio:
                # Peut etre "HH:MM:SS" ou nombre (heures)
                if isinstance(length_audio, (int, float)):
                    heures = float(length_audio)
                    hours = int(heures)
                    minutes = int((heures - hours) * 60)
                    duree_audio = f"{hours}:{minutes:02d}:00"
                else:
                    duree_audio = str(length_audio)

            # Col V (22): Recording Unit Remarks
            recording_remarks = ws.cell(row=row_idx, column=22).value
            if recording_remarks:
                recording_remarks = str(recording_remarks).strip()

            work_orders_data.append({
                'numero_dossier': numero_dossier,
                'date_audience': date_hearing,
                'duree_audio': duree_audio,
                'recording_remarks': recording_remarks,
                'ligne_excel': row_idx
            })

            logger.info(f"   Ligne {row_idx}: {numero_dossier} ({duree_audio})")

            row_idx += 1

        logger.info(f"{len(work_orders_data)} Work Orders detectes dans Excel")

        return work_orders_data

    except Exception as e:
        raise WorkOrderError(f"Erreur parsing Excel multi-Work Orders: {e}")
